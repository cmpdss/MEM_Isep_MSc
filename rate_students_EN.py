import pandas as pd
import re
import time
import logging
import subprocess
import os

from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# ---------- CONFIG ----------
F_STUDENTS = "students.xlsx"       # Your XLS input file
OUTPUT_FILE = "students_grades.xlsx"

ENERGY_SUBJECTS = [
    "Termodinâmica",
    "Mecânica dos Fluidos",
    "Transferência de Calor",
    "Máquinas Térmicas",
    "Máquinas de Fluxo"
]
# -----------------------------------------------------------------------------------------------#
def download_student_photo(driver, student_id, output_dir="photos"):
    """
    Downloads a student's photo from a web page using a Selenium WebDriver instance.
    Args:
        driver (selenium.webdriver.remote.webdriver.WebDriver): 
            The Selenium WebDriver instance used to interact with the web page.
        student_id (str): 
            The unique identifier of the student, used to name the saved photo file.
        output_dir (str, optional): 
            The directory where the photo will be saved. Defaults to "photos".
    Raises:
        Exception: 
            Logs a warning if the photo cannot be extracted or saved.
    Notes:
        - The function assumes the photo is located in an HTML element with the ID 
          "ContentPlaceHolderMain_userpic".
        - If the image source URL starts with "..", it is resolved to an absolute URL 
          based on the ISEP intranet domain.
        - The photo is saved as a PNG file in the specified output directory, with the 
          filename format "{student_id}.png".
        - A new browser tab is opened to load the image, and it is closed after the 
          photo is saved.
    """
    try:
        img = driver.find_element(By.ID, "ContentPlaceHolderMain_userpic")
        img_src = img.get_attribute("src")

        if img_src.startswith(".."):
            img_src = "https://portal.isep.ipp.pt/intranet/" + img_src.replace("..", "")

        #logging.info(f"🔗 Final image URL: {img_src}")

        # Open the image in a new browser tab
        driver.execute_script("window.open();")
        driver.switch_to.window(driver.window_handles[1])
        driver.get(img_src)
        time.sleep(1)  # give it time to load

        # Locate the image and take a screenshot
        img_element = driver.find_element(By.TAG_NAME, "img")
        img_data = img_element.screenshot_as_png

        # Save the screenshot as a PNG
        os.makedirs(output_dir, exist_ok=True)
        save_path = os.path.join(output_dir, f"{student_id}.png")
        with open(save_path, "wb") as f:
            f.write(img_data)

        logging.info(f"📸 Saved photo for {student_id} to: {save_path}")

        # Close tab and switch back
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

    except Exception as e:
        logging.warning(f"⚠️ Could not extract photo for {student_id}: {e}")

# -----------------------------------------------------------------------------------------------#
def extract_students_from_xls(filepath):
    """Extracts student IDs and names from the provided Excel file."""
    logging.info(f"🛠️ Function 'extract_students_from_xls' called")
    logging.info(f"📖 Extracting students from {filepath}")
    if not os.path.exists(filepath):
        logging.error(f"❌ File not found: {filepath}")
        return pd.DataFrame()
    wb = load_workbook(filepath, data_only=True)
    sheet = wb.active

    students = []
    for row in sheet.iter_rows(min_row=1):
        link_cell = row[0]
        name_cell = row[1]

        hyperlink = link_cell.hyperlink.target if link_cell.hyperlink else None
        student_id = link_cell.value  # this is the visible 7-digit student ID
        name = name_cell.value

        if hyperlink:
            match = re.search(r'codeuser=(\d+)', hyperlink)
            codeuser = match.group(1) if match else None
        else:
            codeuser = None

        if codeuser and student_id:
            students.append({
                "student_id": str(student_id),
                "codeuser": codeuser,
                "name": name
            })

    df = pd.DataFrame(students)
    return df

# -----------------------------------------------------------------------------------------------#
def init_driver():
    """Logs into the university portal once and returns the WebDriver session."""
    logging.info(f"🛠️ Function 'init_driver' called")
    USERNAME = "cmi"
    PASSWORD = "3101ev3853XZ"
    try:
        options = webdriver.ChromeOptions()
        options.add_argument("--headless")
        options.add_argument("--disable-gpu")
        options.add_argument("--window-size=1920x1080")
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_argument(
            "user-agent=Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) Chrome/90.0.4430.93 Safari/537.36"
        )
        driver = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver, 15)
        driver.get("https://portal.isep.ipp.pt/intranet/")
        wait.until(
            EC.presence_of_element_located(
                (By.ID, "ContentPlaceHolderMain_txtLoginISEP")
            )
        ).send_keys(USERNAME)
        wait.until(
            EC.presence_of_element_located(
                (By.ID, "ContentPlaceHolderMain_txtPasswordISEP")
            )
        ).send_keys(PASSWORD)
        wait.until(
            EC.element_to_be_clickable(
                (By.ID, "ContentPlaceHolderMain_btLoginISEP")
            )
        ).click()
        time.sleep(2)  # this wait is apparently crucial, don't mess with it
        logging.info("✅ Successfully logged into ISEP portal.")
        return driver
    except Exception as e:
        logging.error(f"❌ Failed to initialize WebDriver: {e}")
        return None

# -----------------------------------------------------------------------------------------------#
def get_all_grades(student_id, codeuser, student_name, driver, idx, total):
    """
    Extracts all grades for a specific student from the ISEP portal and calculates 
    course and energy-related subject averages.
    Args:
        student_id (str): The unique identifier for the student.
        codeuser (str): The codeuser associated with the student in the ISEP portal.
        student_name (str): The full name of the student.
        driver (selenium.webdriver): The Selenium WebDriver instance used for web automation.
        idx (int): The current index of the student being processed.
        total (int): The total number of students being processed.
    Returns:
        tuple: A tuple containing:
            - grades (list): A list of dictionaries where each dictionary represents a grade entry 
              with the following keys: "student_id", "name", "subject", "ects", "grade", "date".
            - course_avg (float or None): The weighted average grade for the entire course, or None if not calculable.
            - energy_avg (float or None): The weighted average grade for energy-related subjects, or None if not calculable.
    Notes:
        - The function navigates to the student's page on the ISEP portal, downloads their photo, 
          and extracts grades from semester tables.
        - It calculates weighted averages for the entire course and for energy-related subjects.
        - Special summary rows for "COURSE AVERAGE" and "ENERGY AVERAGE" are appended to the grades list.
        - If an error occurs during the process, it logs the error and returns empty results.
    """
    grades = []
    url = f"https://portal.isep.ipp.pt/intranet/areapessoal/estudante.aspx?codeuser={codeuser}"
    driver.get(url)

    try:
        wait = WebDriverWait(driver, 15)

        wait.until(
            EC.presence_of_element_located((By.XPATH, f"//*[contains(text(), '{codeuser}')]"))
        )
        download_student_photo(driver, student_id)
        #logging.info(f"✅ [{idx}/{total}] Opened student page for {student_name} (codeuser={codeuser})")
        notas_tab = wait.until(
            EC.element_to_be_clickable((By.ID, "ContentPlaceHolderMain_tab_g"))
        )
        notas_tab.click()
        time.sleep(2)

        semester_tables = driver.find_elements(By.XPATH, "//div[@id='tabStudentFile']//table")
        logging.info(f"📚 Found {len(semester_tables)-1} semester tables.")

        for index, table in enumerate(semester_tables[1:], start=2):
            rows = table.find_elements(By.TAG_NAME, "tr")[1:]  # Skip header
            for row in rows:
                cols = row.find_elements(By.TAG_NAME, "td")
                if len(cols) < 4:
                    continue

                subject = cols[0].text.strip()
                ects = cols[1].text.strip()
                grade = cols[2].text.strip()
                date = cols[3].text.strip()

                if subject == "" or subject.lower() in ["ects", "nota", "data", "unidade curricular"]:
                    continue

                grades.append({
                    "student_id": student_id,
                    "name": student_name,
                    "subject": subject,
                    "ects": ects,
                    "grade": grade,
                    "date": date
                })

        # ➕ Convert to DataFrame to calculate stats
        df = pd.DataFrame(grades)
        df["ects"] = pd.to_numeric(df["ects"], errors="coerce")
        df["grade"] = pd.to_numeric(df["grade"], errors="coerce")
        df = df.dropna(subset=["ects", "grade"])

        # 📊 Course Average
        total_ects = df["ects"].sum()
        course_avg = (df["grade"] * df["ects"]).sum() / total_ects if total_ects > 0 else None

        # ⚡ Energy Subjects Average
        energy_subjects = [
            "Termodinâmica",
            "Mecânica dos Fluidos",
            "Transferência de Calor",
            "Máquinas Térmicas",
            "Máquinas de Fluxo"
        ]
        df_energy = df[df["subject"].isin(ENERGY_SUBJECTS)]
        energy_ects = df_energy["ects"].sum()
        energy_avg = (df_energy["grade"] * df_energy["ects"]).sum() / energy_ects if energy_ects > 0 else None

        # ➕ Append to grades as special summary rows
        grades.append({
            "student_id": student_id,
            "name": student_name,
            "subject": "COURSE AVERAGE",
            "ects": "",
            "grade": round(course_avg, 2) if course_avg else "N/A",
            "date": ""
        })

        grades.append({
            "student_id": student_id,
            "name": student_name,
            "subject": "ENERGY AVERAGE",
            "ects": "",
            "grade": round(energy_avg, 2) if energy_avg else "N/A",
            "date": ""
        })

        logging.info(f"📚 Média de curso   : {course_avg:.2f}" if course_avg else "📚 Média de curso   : N/A")
        logging.info(f"📚 Média de ENERGIA : {energy_avg:.2f}" if energy_avg else "📚 Média de ENERGIA : N/A")

        return grades, round(course_avg, 2) if course_avg else None, round(energy_avg, 2) if energy_avg else None

    except Exception as e:
        logging.error(f"❌ Failed to extract grades for {student_name} ({student_id})")
        return [], None, None

# -----------------------------------------------------------------------------------------------#
def process_all_students(filepath):
    """
    Processes student data from an Excel file, retrieves their grades, and generates summaries.

    Args:
        filepath (str): The file path to the Excel file containing student data.

    Returns:
        tuple: A tuple containing:
            - all_grade_records (list): A list of all grade records for the students.
            - summary_rows (list): A list of dictionaries summarizing each student's global and energy averages.
            - missing_students (list): A list of dictionaries for students with no grade data.
    """
    df_students = extract_students_from_xls(filepath)
    driver = init_driver()
    if driver is None:
        return [], []

    all_grade_records = []
    summary_rows = []  # One row per student
    missing_students = []  # To collect students without academic data

    for idx, row in df_students.iterrows():
        student_id = row["student_id"]
        codeuser = row["codeuser"]
        name = row["name"]
        logging.info("-" * 80)
        logging.info(f"🔍 {idx+1} / {len(df_students)} Processing {name} ({student_id})")
        grades, course_avg, energy_avg = get_all_grades(student_id, codeuser, name, driver, idx + 1, len(df_students))
        all_grade_records.extend(grades)

        summary_rows.append({
            "student_id": student_id,
            "full_name": name,
            "global_average": course_avg,
            "energy_average": energy_avg
        })

        # Track students with no grade data
        if course_avg is None and energy_avg is None:
            missing_students.append({
                "student_id": student_id,
                "full_name": name
            })

        time.sleep(1)

    driver.quit()
    return all_grade_records, summary_rows, missing_students

# -----------------------------------------------------------------------------------------------#
def export_grades_to_excel(grade_records, output_file):
    """
    Exports a list of grade records to an Excel file.
    Args:
        grade_records (list of dict): A list where each dictionary represents a student's grade record.
        output_file (str): The file path where the Excel file will be saved.
    Returns:
        None
    Side Effects:
       
        - Creates an Excel file at the specified output path.
        - Logs a message indicating the success of the export operation.
    """
    df = pd.DataFrame(grade_records)
    df.to_excel(output_file, index=False)
    logging.info(f"✅ Grades exported to: {output_file}")

# -----------------------------------------------------------------------------------------------#
def export_energy_shortlist(summary_rows, output_file="shortlist_energy_top20.xlsx"):
    df = pd.DataFrame(summary_rows)
    df = df.dropna(subset=["energy_average"])  # Keep only students with energy grades
    df = df.sort_values(by="energy_average", ascending=False)

    top_n = max(1, int(len(df) * 1.0))  # Top 100%
    shortlist = df.head(top_n)

    shortlist.to_excel(output_file, index=False)
    logging.info(f"🏆 Shortlist of top {top_n} students saved to {output_file}")

# -----------------------------------------------------------------------------------------------#
def export_to_latex(results, summary, missing_students, photo_dir="photos", tex_file="students_report.tex"):
    """
    Generates a LaTeX report for students, including their academic performance and photos.
    This function creates a LaTeX document that includes detailed information about students,
    such as their grades, averages, and photos. It also lists students with unknown academic
    performance in a separate section.
    Args:
        results (list of dict): A list of dictionaries containing student grades. Each dictionary
            should include keys like 'student_id', 'subject', 'grade', and 'date'.
        summary (list of dict): A list of dictionaries summarizing student performance. Each
            dictionary should include keys like 'student_id', 'full_name', 'global_average',
            and 'energy_average'.
        missing_students (list of dict): A list of dictionaries for students with unknown
            academic performance. Each dictionary should include keys like 'student_id' and 'full_name'.
        photo_dir (str, optional): The directory containing student photos. Defaults to "photos".
        tex_file (str, optional): The output LaTeX file path. Defaults to "students_report.tex".
    Returns:
        None: The function writes the LaTeX content to the specified file.
    Raises:
        KeyError: If required keys are missing in the input dictionaries.
        FileNotFoundError: If a photo file for a student is not found in the specified directory.
    Notes:
        - The LaTeX document uses specific packages such as `fontspec`, `graphicx`, and `geometry`.
        - Students are sorted by their energy average in descending order.
        - Ensure that the `photo_dir` contains images named as `{student_id}.png` for each student.
    """
    df_grades = pd.DataFrame(results)
    df_summary = pd.DataFrame(summary)

    # Ensure sorting by descending energy average
    df_summary = df_summary.dropna(subset=["energy_average"]).sort_values(by="energy_average", ascending=False)

    latex_lines = [
        r"\documentclass[10pt]{article}",
        r"\usepackage{fontspec}",
        r"\setmainfont{Arial}",
        r"\usepackage{graphicx}",
        r"\usepackage[margin=1in]{geometry}",
        r"\usepackage{float}",
        r"\usepackage{titlesec}",
        r"\usepackage{enumitem}",
        r"\setlist[itemize]{noitemsep, topsep=0pt}",
        r"\begin{document}",
        r" ",
        r"\centering{\bf{\Large{Lista de alunos do 1º ano do MEM, Ramo de Energia, 2024/2025}}}",
        r"\vspace{1em}",
        r"\hrule",
        r"\vspace{1em}"
    ]

    for _, row in df_summary.iterrows():
        student_id = row["student_id"]
        name = row["full_name"]
        global_avg = row["global_average"]
        energy_avg = row["energy_average"]

        student_grades = df_grades[
            (df_grades["student_id"] == row["student_id"]) &
            (df_grades["subject"].isin(ENERGY_SUBJECTS))
        ][["subject", "grade", "date"]]

        latex_lines.extend([
            r"\noindent\begin{minipage}{0.25\textwidth}",
            rf"\includegraphics[width=\linewidth]{{{photo_dir}/{student_id}.png}}",
            r"\end{minipage}%",
            r"\hspace{0.2cm}",
            r"\hfill\begin{minipage}{0.60\textwidth}",
            rf"\textbf{{Número}}: {student_id} \\",
            rf"\textbf{{Nome}}: {name} \\",
            r"\vspace{0.5em}",
            r"\begin{itemize}",
        ])

        for _, g in student_grades.iterrows():
            if 'date' in g and g['date']:
                latex_lines.append(rf"\item {g['subject']}: {g['grade']} [{g['date']}]")
            else:
                latex_lines.append(rf"\item {g['subject']}: {g['grade']}")

        latex_lines.extend([
            r"\vspace{0.5em}",
            rf"\item \textbf{{MÉDIA DE CURSO}}   : \textbf{{{global_avg:.2f}}}"
            rf"\item \textbf{{MÉDIA UCs ENERGIA}}: \textbf{{{energy_avg:.2f}}}",
            r"\end{itemize}",
            r"\end{minipage}",
            r"\vspace{1em}",
            r"\hrule",
            r"\vspace{1em}"
        ])

    if missing_students:
        latex_lines.append(r"\newpage")
        latex_lines.append(r"\centering{\bf{\Large{Estudantes que não vieram da LEM}}}")
        latex_lines.append(r"\vspace{1em}")
        latex_lines.append(r"\hrule")
        latex_lines.append(r"\vspace{1em}")

        for student in missing_students:
            student_id = student["student_id"]
            name = student["full_name"]

            latex_lines.extend([
                r"\noindent\begin{minipage}{0.25\textwidth}",
                rf"\includegraphics[width=\linewidth]{{{photo_dir}/{student_id}.png}}",
                r"\end{minipage}%",
                r"\hspace{0.2cm}",
                r"\hfill\begin{minipage}{0.70\textwidth}",
                rf"\textbf{{Número}}: {student_id} \\",
                rf"\textbf{{Nome}}: {name}",
                r" ",
                r"\vspace{1em}",
                r"Aluno/a não frequentou a Licenciatura em Engenharia Mecânica.\\",
                r"Desempenho académico desconhecido.",
                r"\end{minipage}",
                r"\vspace{1em}",
                r"\hrule",
                r"\vspace{1em}"
            ])

    latex_lines.append(r"\end{document}")

    with open(tex_file, "w", encoding="utf-8") as f:
        f.write("\n".join(latex_lines))

    logging.info(f"📄 LaTeX file written to: {tex_file}")

# -----------------------------------------------------------------------------------------------#
def compile_pdf_and_open(tex_file="students_report.tex"):
    """
    Compiles a LaTeX file into a PDF using the XeLaTeX engine and opens the resulting PDF.
    Args:
        tex_file (str): The name of the LaTeX file to compile. Defaults to "students_report.tex".
    Logs:
        - Logs an informational message if the compilation is successful.
        - Logs an error message if the compilation fails.
    Raises:
        subprocess.CalledProcessError: If the XeLaTeX compilation or opening the PDF fails.
    """
    try:
        subprocess.run(["xelatex", tex_file], check=True)
        logging.info(f"📄 Compiled {tex_file} to PDF using xelatex.")
        pdf_file = os.path.splitext(tex_file)[0] + ".pdf"
        subprocess.run(["open", pdf_file], check=True)
    except subprocess.CalledProcessError as e:
        logging.error(f"❌ Failed to compile LaTeX: {e}")

# -----------------------------------------------------------------------------------------------#
# Main function
# -----------------------------------------------------------------------------------------------#        
if __name__ == "__main__":
    # Set up logging
    LOG_FILE = "rate_students_EN.log"
    if os.path.exists(LOG_FILE):
        os.remove(LOG_FILE)

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(LOG_FILE),  # ✅ Log to File
            logging.StreamHandler(),  # ✅ Log to Terminal
        ],
        force=True,  # ✅ Force reconfiguration of logging
    )

    logging.info("🚀 Starting full-grade extraction...")
    results, summary, missing_students = process_all_students(F_STUDENTS)

    export_grades_to_excel(results, OUTPUT_FILE)
    export_to_latex(results, summary, missing_students)
    compile_pdf_and_open("students_report.tex")
    #export_energy_shortlist(summary)
    logging.info("🏁 Finished.")